# -*- coding: utf-8 -*-
"""
Created on Sun Feb 26 18:25:09 2023

@author: Sarah
"""
import tkinter as tk
from tkinter import *
from tkinter import ttk
from openpyxl import load_workbook
import pandas as pd
import webbrowser

global all_groups #holds the categories for each life
global all_groups_indexes #holds the index of the start of each sub group
global all_items #holds the crafting items list
global all_inventory #holds the inventory items list
global all_locations #holds the location counts
global curr_screen
global text_frame
curr_screen = 0
global curr_subscreen
curr_subscreen = 0
all_groups = []
all_items = []
all_groups_indexes = []
global wb #excel workbook Shopping.xlsx
global text_scroll
global button_manager
global entry_arr
entry_arr = []
global shopping_list
shopping_list = []
global reset_count 
reset_count = 0
global wentShopping
wentShopping = 0

class CraftingItem:
    def __init__(self, name, ingredients, counts, total):
        self.name = name
        self.ingredients = ingredients
        self.counts = counts
        self.total = total

class InventoryItem:
    def __init__(self, name, cost, locations, link):
        self.name = name
        self.cost = cost
        self.locations = locations
        self.link = link
        
class LocationList:
    def __init__(self, name, listOfItems, listOfInvIndexes):
        self.name = name
        self.listOfItems = listOfItems
        self.listOfInvIndexes = listOfInvIndexes

class ShoppingItem:
    def __init__(self, name, count):
        self.name = name
        self.count = count
        
def resetResetButton():
    global button_manager
    global reset_count
    reset_count = 0
    button_manager[10].configure(text="Reset Everything?")

def reset():
    global reset_count
    global button_manager
    global shopping_list
    global entry_arr
    
    if reset_count == 0:
        reset_count = 1
        button_manager[10].configure(text="Are You Sure?")
    else:
        resetResetButton()
        shopping_list = []
        for i in range(0,5):
            for j in range(len(all_items[i])):
                all_items[i][j].total = 0
                
        for i in range(len(entry_arr)):
            entry_arr[i].delete(0,'end')
            entry_arr[i].insert(0,0)
        if wentShopping == 1:
            goShopping()
            
    

def resetOnlyItems():
    global entry_arr
    global all_items
    global wentShopping 
    global no_location
    wentShopping = 0
    
    global all_locations
    global shopping_list
    
    tot = 0
    for i in range(len(all_locations)):
        for j in range(len(all_locations[i].listOfItems)):
                for k in range(len(shopping_list)):
                    if shopping_list[k].name == all_locations[i].listOfItems[j].name and entry_arr[tot].get().isnumeric():
                        shopping_list[k].count = int(entry_arr[tot].get())
                        if shopping_list[k].count == 0:
                                shopping_list.pop(k)
                        break
                        
                tot = tot + 1
                
    for i in range(len(no_location)):
        for k in range(len(shopping_list)):
            if shopping_list[k].name == no_location[i].name and entry_arr[tot].get().isnumeric():
                shopping_list[k].count = int(entry_arr[tot].get())
                if shopping_list[k].count == 0:
                    shopping_list.pop(k)
                break
                    
        tot = tot + 1

    for i in range(0,5):
        for j in range(len(all_items[i])):
            all_items[i][j].total = 0
                
    for i in range(len(entry_arr)):
          entry_arr[i].delete(0,'end')
          entry_arr[i].insert(0,0)
          
         
          

def link(v):
    webbrowser.open_new(v)
    
def refreshShopping():
    global all_locations
    
    for i in range(len(all_locations)):
        all_locations[i].listOfItems = []
        all_locations[i].listOfInvIndexes = []
 

def goShopping(*args):
    updateShoppingList()
    resetResetButton()
    global text_scroll
    global text_frame
    global shopping_list
    global entry_arr
    global all_inventory
    global all_locations
    global wentShopping
    global no_location
    
    #update elements and erase if 0
    if(wentShopping == 1):
        tot = 0
        for i in range(len(all_locations)):
            for j in range(len(all_locations[i].listOfItems)):
                    for k in range(len(shopping_list)):
                        if shopping_list[k].name == all_locations[i].listOfItems[j].name and entry_arr[tot].get().isnumeric():
                            shopping_list[k].count = int(entry_arr[tot].get())
                            if shopping_list[k].count == 0:
                                shopping_list.pop(k)
                            break
                            
                    tot = tot + 1
                    
        for i in range(len(no_location)):
                for k in range(len(shopping_list)):
                    if shopping_list[k].name == no_location[i].name and entry_arr[tot].get().isnumeric():
                        shopping_list[k].count = int(entry_arr[tot].get())
                        if shopping_list[k].count == 0:
                            shopping_list.pop(k)
                        break
                            
                tot = tot + 1
    
    wentShopping = 1
    entry_arr = []
    no_location = []
    
    refreshShopping()
    
    ##destroy what was already there
    text_frame.destroy()
    text_frame = Frame(holdSelf)
    text_frame.pack_propagate(0)
    text_frame.config(height=650, width=650)
    text_frame.pack()
    text_scroll = Scrollbar(text_frame, "Text")
    
    #iterate each shopping_list item and place them into all_locations lists based on first index
    for i in range(len(shopping_list)):
        ind = -1
        for j in range(len(all_inventory)):
            if shopping_list[i].name == all_inventory[j].name:
                ind = j
                break
        if ind != -1:
            #append shopping list name and count to the location file where they are located
            all_locations[all_inventory[ind].locations[0]].listOfItems.append(shopping_list[i]) 
            all_locations[all_inventory[ind].locations[0]].listOfInvIndexes.append(ind) 
        else:
            no_location.append(shopping_list[i])
        
     #iterate through each shopping_list item and move it into a different list if later location available
    for i in range(len(shopping_list)):
        ind = -1
        for j in range(len(all_inventory)):
            if shopping_list[i].name == all_inventory[j].name:
                ind = j
                break
        if ind != -1:
            found = -1
            for j in range(len(all_inventory[ind].locations)-1):
                temp = len(all_inventory[ind].locations)-1-j
                if len(all_locations[all_inventory[ind].locations[temp]].listOfItems) > 0:
                    all_locations[all_inventory[ind].locations[temp]].listOfItems.append(shopping_list[i])
                    all_locations[all_inventory[ind].locations[temp]].listOfInvIndexes.append(ind)
                    
                    for k in range(len(all_locations[all_inventory[ind].locations[0]].listOfItems)):
                        if all_locations[all_inventory[ind].locations[0]].listOfItems[k].name == all_inventory[ind].name:
                            all_locations[all_inventory[ind].locations[0]].listOfItems.pop(k)
                            all_locations[all_inventory[ind].locations[0]].listOfInvIndexes.pop(k)
                            found = 1
                            break
                if found == 1:
                    found = -1
                    break
        
    tot = 0
    for i in range(len(all_locations)):
        for j in range(len(all_locations[i].listOfItems)):
            temp_button = Button(text_scroll.frame, width=2, text="-", command=lambda v=tot: decrement(v))
            temp_button.grid(row = tot+2, column=0, sticky = W, pady = 2)
            
            entry_arr.append(tk.Entry(text_scroll.frame, width = 3))
            entry_arr[tot].insert(0,all_locations[i].listOfItems[j].count)
            entry_arr[tot].grid(row = tot+2, column=1, sticky = W, pady = 2)
            
            temp_button = Button(text_scroll.frame, width=2, text="+", command=lambda v=tot: increment(v))
            temp_button.grid(row = tot+2, column=2, sticky = W, pady = 2)
            
            temp_button = Button(text_scroll.frame, text=all_locations[i].listOfItems[j].name, command=lambda v=all_inventory[all_locations[i].listOfInvIndexes[j]].link: link(v))
            temp_button.grid(row = tot+2, column = 3)
            
            tk.Label(text_scroll.frame, text=all_locations[i].name).grid(row=tot+2, column=4, sticky='nw')
            
            tot = tot+1
    
    for i in range(len(no_location)):
            temp_button = Button(text_scroll.frame, width=2, text="-", command=lambda v=tot: decrement(v))
            temp_button.grid(row = tot+2, column=0, sticky = W, pady = 2)
            
            entry_arr.append(tk.Entry(text_scroll.frame, width = 3))
            entry_arr[tot].insert(0,no_location[i].count)
            entry_arr[tot].grid(row = tot+2, column=1, sticky = W, pady = 2)
            
            temp_button = Button(text_scroll.frame, width=2, text="+", command=lambda v=tot: increment(v))
            temp_button.grid(row = tot+2, column=2, sticky = W, pady = 2)
            
            search_link = "https://fantasy-life.fandom.com/wiki/Special:Search?query=" + no_location[i].name
            temp_button = Button(text_scroll.frame, text=no_location[i].name, command=lambda v=search_link: link(v))
            temp_button.grid(row = tot+2, column = 3)
            
            tk.Label(text_scroll.frame, text="No Purchase Location").grid(row=tot+2, column=4, sticky='nw')
            
            tot = tot + 1
    
    
    
        
def updateShoppingList():
    global curr_screen
    global curr_subscreen
    global entry_arr
    global all_groups_indexes
    global all_items
    global shopping_list
    
    #determine start and end indexes for each subsection
    temp_index_start = all_groups_indexes[curr_screen][curr_subscreen]
    
    if curr_subscreen != 3:
        temp_index_end = all_groups_indexes[curr_screen][curr_subscreen+1]
    else:
        temp_index_end = len(all_items[curr_screen])
        
    #add items to the shopping list
    for i in range(temp_index_start, temp_index_end):
        if len(entry_arr) == temp_index_end - temp_index_start:
            if entry_arr[i-temp_index_start].get().isnumeric():
                if int(entry_arr[i-temp_index_start].get()) > 0:
                    all_items[curr_screen][i].total = int(entry_arr[i-temp_index_start].get())
                    for j in range(len(all_items[curr_screen][i].ingredients)):
                        checker = -1
                        for k in range(len(shopping_list)):
                            if all_items[curr_screen][i].ingredients[j] == shopping_list[k].name:
                                checker = k
                                break
                        if checker != -1: #if the item already in shopping list, update total
                            shopping_list[checker].count = shopping_list[checker].count + all_items[curr_screen][i].counts[j] * int(entry_arr[i-temp_index_start].get())
                        else: #if the item not already in shopping list, add it
                            shopping_list.append(ShoppingItem(all_items[curr_screen][i].ingredients[j], all_items[curr_screen][i].counts[j] * int(entry_arr[i-temp_index_start].get())))
            

        
def decrement(v):
    global entry_arr
    resetResetButton()
    
    if entry_arr[v].get().isnumeric():
        num = int(entry_arr[v].get())
        if num > 0:
            num = num - 1
            entry_arr[v].delete(0,'end')
            entry_arr[v].insert(0,num)
    
def increment(v):
    global entry_arr
    resetResetButton()
    
    if entry_arr[v].get().isnumeric():
        num = int(entry_arr[v].get())
        num = num + 1
        entry_arr[v].delete(0,'end')
        entry_arr[v].insert(0,num)
        
    

def updateShoppingText():
    global curr_screen
    global curr_subscreen
    global all_groups_indexes
    global text_frame
    global text_scroll
    global entry_arr
    
    temp_index_start = 0
    temp_index_end = 0
    
    ##destroy what was already there
    text_frame.destroy()
    text_frame = Frame(holdSelf)
    text_frame.pack_propagate(0)
    text_frame.config(height=650, width=650)
    text_frame.pack()
    text_scroll = Scrollbar(text_frame, "Text")
    
    entry_arr = []
    
    #determine start and end indexes for each subsection
    temp_index_start = all_groups_indexes[curr_screen][curr_subscreen]
    if curr_subscreen != 3:
        temp_index_end = all_groups_indexes[curr_screen][curr_subscreen+1]
    else:
        temp_index_end = len(all_items[curr_screen])
        
    for i in range(temp_index_start, temp_index_end):
        temp_item_string = ""
        for j in range(len(all_items[curr_screen][i].ingredients)):
            if j > 0:
                temp_item_string = temp_item_string + ", "
            temp_item_string = temp_item_string + all_items[curr_screen][i].ingredients[j] + " x" + str(all_items[curr_screen][i].counts[j])
        
        tk.Label(text_scroll.frame, text=all_items[curr_screen][i].name).grid(row=i+2-temp_index_start, column=3, sticky='nw')
        tk.Label(text_scroll.frame, text=temp_item_string).grid(row=i+2-temp_index_start, column=4, sticky='nw')
        temp_button = Button(text_scroll.frame, width=2, text="-", command=lambda v=i-temp_index_start: decrement(v))
        temp_button.grid(row = i+2-temp_index_start, column=0, sticky = W, pady = 2)
        entry_arr.append(tk.Entry(text_scroll.frame, width = 3))
        entry_arr[i-temp_index_start].insert(0,all_items[curr_screen][i].total)
        entry_arr[i-temp_index_start].grid(row = i+2-temp_index_start, column=1, sticky = W, pady = 2)
        temp_button = Button(text_scroll.frame, width=2, text="+", command=lambda v=i-temp_index_start: increment(v))
        temp_button.grid(row = i+2-temp_index_start, column=2, sticky = W, pady = 2)
    
    

def loadScreen(v, *args):
    updateShoppingList()
    resetResetButton()
    
    global wentShopping
    if wentShopping == 1:
        resetOnlyItems()
    
    global curr_screen
    global curr_subscreen
    
    old_screen = curr_screen
    curr_screen = v
    
    old_subscreen = curr_subscreen
    curr_subscreen = 0
    
    #update bold-ness of buttons
    global button_manager
    button_manager[old_screen].configure(font=('TkDefaultFont', 9, 'normal'))
    button_manager[v].configure(font=('TkDefaultFont', 9, 'bold'))
    
    button_manager[old_subscreen+6].configure(font=('TkDefaultFont', 9, 'normal'))
    button_manager[6].configure(font=('TkDefaultFont', 9, 'bold'))
    
    button_manager[6].configure(text=all_groups[curr_screen][0])
    button_manager[7].configure(text=all_groups[curr_screen][1])
    button_manager[8].configure(text=all_groups[curr_screen][2])
    button_manager[9].configure(text=all_groups[curr_screen][3])
    
    
    updateShoppingText()
    
    
def loadSubScreen(v, *args):
    updateShoppingList()
    resetResetButton()
    
    global wentShopping
    if wentShopping == 1:
        resetOnlyItems()
    
    global curr_subscreen
    
    old_subscreen = curr_subscreen
    curr_subscreen = v - 6
    
    #update bold-ness of buttons
    global button_manager
    button_manager[old_subscreen+6].configure(font=('TkDefaultFont', 9, 'normal'))
    button_manager[v].configure(font=('TkDefaultFont', 9, 'bold'))
    
    updateShoppingText()


def loopTheSheets(sheetName, j):
    global all_groups
    global all_items
    all_groups.append([])
    all_groups_indexes.append([])
    all_items.append([])
    i = 1
    temp_ingredients = []
    temp_counts = []
    while(wb[sheetName].cell(row=i, column=1).value != None or wb[sheetName].cell(row=i, column=2).value != None):
        if wb[sheetName].cell(row=i, column=2).value == None:
            all_groups[j].append(wb[sheetName].cell(row=i, column=1).value)
            i = i + 1
            all_groups_indexes[j].append(len(all_items[j]))
        elif wb[sheetName].cell(row=i, column=1).value != None:
            temp_ingredients.append(wb[sheetName].cell(row=i, column=2).value)
            temp_counts.append(wb[sheetName].cell(row=i, column=3).value)
            temp_name = wb[sheetName].cell(row=i, column=1).value
            i = i + 1 
            if wb[sheetName].cell(row=i, column=1).value == None and wb[sheetName].cell(row=i, column=2).value != None:
                temp_ingredients.append(wb[sheetName].cell(row=i, column=2).value)
                temp_counts.append(wb[sheetName].cell(row=i, column=3).value)
                i = i + 1 
                if wb[sheetName].cell(row=i, column=1).value == None and wb[sheetName].cell(row=i, column=2).value != None:
                    temp_ingredients.append(wb[sheetName].cell(row=i, column=2).value)
                    temp_counts.append(wb[sheetName].cell(row=i, column=3).value)
                    i = i + 1 
            temp_craftingitem = CraftingItem(temp_name, temp_ingredients, temp_counts, 0)
            temp_ingredients = []
            temp_counts = []
            all_items[j].append(temp_craftingitem)
            
   


def loadDataFromExcel(wb):
    global all_locations
    global all_inventory
    all_locations = []
    all_inventory = []
    loopTheSheets('Alchemist', 0)
    loopTheSheets('Cook', 1)
    loopTheSheets('Blacksmith', 2)
    loopTheSheets('Carpenter', 3)
    loopTheSheets('Tailor', 4)
    
    #initialize locations list
    i = 2
    while(wb['Shops'].cell(row=1, column=i).value != None):
        all_locations.append(LocationList(wb['Shops'].cell(row=1, column=i).value,[],[]))
        i = i+1
        
    #load in the shopping list
    i=2
    while(wb['Shops'].cell(row=i, column=1).value != None):
        temp_name = wb['Shops'].cell(row=i, column=1).value
        temp_locations = []
        j=0
        for j in range(len(all_locations)):
            if wb['Shops'].cell(row=i, column=j+2).value != None:
                temp_cost = wb['Shops'].cell(row=i, column=j+2).value
                temp_locations.append(j)
        all_inventory.append(InventoryItem(temp_name, temp_cost, temp_locations, wb['Shops'].cell(row=i, column=1).hyperlink.target))
        i = i+1
        
     
    
def on_mousewheel(event):
    global text_scroll
    shift = (event.state & 0x1) != 0
    scroll = -1 if event.delta > 0 else 1
    if shift:
        text_scroll.canvas.xview_scroll(scroll, "units")
    else:
        text_scroll.canvas.yview_scroll(scroll, "units")
        

class Scrollbar(tk.Frame):
    def __init__(self, root, name_of_type):

        
        tk.Frame.__init__(self, root)
        self.canvas = tk.Canvas(root, borderwidth=0)
        self.frame = tk.Frame(self.canvas)
        self.vsb = tk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)
        if name_of_type != "Map":
            self.hsb = tk.Scrollbar(root, orient="horizontal", command=self.canvas.xview)
            self.canvas.configure(xscrollcommand=self.hsb.set)
            self.hsb.pack(side="bottom", fill="x")
            
        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window((4,4), window=self.frame, anchor="nw", 
                                  tags="self.frame")
        
        self.frame.bind("<Configure>", self.onFrameConfigure) 
    
    def onFrameConfigure(self, event):
        '''Reset the scroll region to encompass the inner frame'''
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        

class Window(Scrollbar):
    def __init__(self, master=None):
        Frame.__init__(self, master)               
        self.master = master
        self.init_window()
        
        
    def init_window(self):
        # allowing the widget to take the full space of the root window
        self.pack(fill=BOTH, expand=1)
        global holdSelf
        holdSelf = self
        global titleLocation
        titleLocation = self.master
        titleLocation.title("Fantasy Life Shopping Tool")
        
        button_frame = Frame(self)
        button_frame.pack()
        
        global text_frame
        text_frame = Frame(self)
        text_frame.pack_propagate(0)
        text_frame.config(width=650, height=650)
        text_frame.pack()
        
        global text_scroll
        text_scroll = Scrollbar(text_frame, "Text")
        
        # creating buttons
        global button_manager
        button_manager = [None]*11
        
        button_manager[0] = Button(button_frame, width=13, text="Alchemist", font = ('TkDefaultFont', 9, 'bold'), command=lambda v=0: loadScreen(v))
        button_manager[1] = Button(button_frame, width=13, text="Cook", command=lambda v=1: loadScreen(v))
        button_manager[2] = Button(button_frame, width=13, text="Blacksmith", command=lambda v=2: loadScreen(v))
        button_manager[3] = Button(button_frame, width=13, text="Carpenter", command=lambda v=3: loadScreen(v))
        button_manager[4] = Button(button_frame, width=13, text="Tailor", command=lambda v=4: loadScreen(v))
        button_manager[5] = Button(button_frame, width=13, text="Go Shopping!", command=lambda: goShopping())
        
        button_manager[6] = Button(button_frame, width=13, text=all_groups[0][0], font = ('TkDefaultFont', 9, 'bold'), command=lambda v=6: loadSubScreen(v))
        button_manager[7] = Button(button_frame, width=13, text=all_groups[0][1], command=lambda v=7: loadSubScreen(v))
        button_manager[8] = Button(button_frame, width=13, text=all_groups[0][2], command=lambda v=8: loadSubScreen(v))
        button_manager[9] = Button(button_frame, width=13, text=all_groups[0][3], command=lambda v=9: loadSubScreen(v))
        
        button_manager[10] = Button(button_frame, width=13, text="Reset Everything?", command=lambda: reset())
        
        # placing buttons
        for i in range(6):
            button_manager[i].grid(row = 0, column = i)
            
        for i in range(6,10):
            button_manager[i].grid(row = 1, column = i-6)
            
        button_manager[10].grid(row = 1, column = 5)
        
        
    def client_exit(self):
        wb.close()
        exit() 


###############MAIN##################
        

   

try:
    global wb  
    wb = load_workbook('Shopping.xlsx')
except FileNotFoundError:
    print("ERROR")
    sys.exit()
    
loadDataFromExcel(wb)
wb.close()

root = Tk()

#size of the window
root.geometry("650x650")
root.bind_all("<MouseWheel>", on_mousewheel)
root.bind_all("<Button-4>", on_mousewheel)
root.bind_all("<Button-5>", on_mousewheel)
app = Window(root)

loadScreen(0)

root.mainloop()
